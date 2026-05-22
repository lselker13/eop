#!/usr/bin/env python3
"""
Download COD-AB p-code tables from OCHA Humanitarian Data Exchange (HDX)
for each country in the EOP survey.

Source: HDX package cod-ab-{iso3} (OCHA standardised Common Operational
Dataset – Administrative Boundaries). Each package contains one XLSX with
all admin levels. No fallback is attempted for countries without a
cod-ab-{iso3} package.

Survey years are read from:
  /data/eop/compiled_country_data/survey_countries.csv  (field: survey_year)
Only rows with using=True are processed; when a country appears multiple
times the latest survey year is used.

Output per country:
  /data/eop/geo/humdata_tables/{ISO3}/{filename}_{sheet}.csv  (one CSV per XLSX sheet)
  /data/eop/geo/humdata_tables/{ISO3}/readme.txt

A machine-readable manifest summarising all outcomes is written to:
  /data/eop/geo/humdata_tables/manifest.json
"""

import csv
import json
import os
import re
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SURVEY_CSV = "/data/eop/compiled_country_data/survey_countries.csv"
OUTPUT_DIR = Path("/data/eop/geo/humdata_tables")
HDX_PACKAGE_URL = "https://data.humdata.org/api/3/action/package_show"
ACCESS_DATE = date.today().isoformat()

TABULAR_FORMATS = {"csv", "xlsx", "xls"}
REQUEST_DELAY = 0.5  # seconds between API/download calls


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_survey_years() -> dict[str, float]:
    """Return {country_code: survey_year} for using=True rows.

    When a country code appears more than once the latest year is kept,
    so a single target vintage is chosen per country.
    """
    years: dict[str, float] = {}
    with open(SURVEY_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("using", "").strip().lower() != "true":
                continue
            code = row["country_code"].strip()
            try:
                yr = float(row["survey_year"])
            except (ValueError, KeyError):
                continue
            if code not in years or yr > years[code]:
                years[code] = yr
    return years


# ---------------------------------------------------------------------------
# HDX API
# ---------------------------------------------------------------------------

def fetch_cod_ab_package(iso3: str) -> dict | None:
    """Return the cod-ab-{iso3} HDX dataset dict, or None if not found."""
    try:
        r = requests.get(
            HDX_PACKAGE_URL,
            params={"id": f"cod-ab-{iso3.lower()}"},
            timeout=20,
        )
        r.raise_for_status()
        j = r.json()
        return j["result"] if j.get("success") else None
    except Exception as e:
        print(f"  API error: {e}")
        return None


def parse_dataset_date_start(dataset: dict) -> date | None:
    """Extract the start date from the HDX dataset_date range string.

    HDX stores this as '[YYYY-MM-DDT... TO YYYY-MM-DDT...]'. The start date
    is the earliest date for which the boundaries are considered valid.
    Falls back to metadata_created if the range is absent.
    """
    raw = dataset.get("dataset_date", "")
    m = re.search(r"\[(\d{4}-\d{2}-\d{2})", raw)
    if m:
        try:
            return datetime.fromisoformat(m.group(1)).date()
        except ValueError:
            pass
    raw2 = dataset.get("metadata_created", "")
    if raw2:
        try:
            return datetime.fromisoformat(raw2[:10]).date()
        except ValueError:
            pass
    return None


def select_tabular_resources(dataset: dict, survey_year: float) -> list[dict]:
    """Return tabular (CSV/XLSX/XLS) resources from the dataset.

    For the standardised cod-ab packages there is exactly one XLSX; this
    function returns it directly. For older-format packages that carry
    separate per-level CSVs, it groups by admin level and selects the
    resource whose last_modified year is closest to survey_year.
    """
    tabular = [
        r for r in dataset.get("resources", [])
        if r.get("format", "").lower() in TABULAR_FORMATS
    ]
    if not tabular:
        return []
    if len(tabular) == 1:
        return tabular

    # Multiple resources: group by admin level, pick nearest per level.
    level_pat = re.compile(r"adm[_\-]?(\d)", re.IGNORECASE)

    def res_year(res: dict) -> int:
        for field in ("last_modified", "created"):
            val = res.get(field, "")
            if val:
                try:
                    return datetime.fromisoformat(val[:10]).year
                except ValueError:
                    pass
        ds_start = parse_dataset_date_start(dataset)
        return ds_start.year if ds_start else date.today().year

    groups: dict[str, list[dict]] = {}
    no_level: list[dict] = []
    for res in tabular:
        m = level_pat.search(res.get("name", ""))
        key = m.group(0).lower().replace("-", "").replace("_", "") if m else None
        (groups.setdefault(key, []) if key else no_level).append(res)

    selected: list[dict] = []
    for items in list(groups.values()) + ([no_level] if no_level else []):
        best = min(items, key=lambda r: abs(res_year(r) - int(survey_year)))
        selected.append(best)
    return selected


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_file(url: str, dest: Path) -> int:
    """Stream-download url to dest; return file size in bytes."""
    r = requests.get(url, stream=True, timeout=180)
    r.raise_for_status()
    with open(dest, "wb") as f:
        for chunk in r.iter_content(chunk_size=65536):
            f.write(chunk)
    return dest.stat().st_size


def safe_filename(res: dict) -> str:
    """Derive a safe local filename from the resource name/URL."""
    raw = res.get("name", "") or res.get("url", "").split("/")[-1] or "file"
    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in raw)
    fmt = res.get("format", "xlsx").lower()
    if not any(safe.lower().endswith(f".{f}") for f in TABULAR_FORMATS):
        safe += f".{fmt}"
    return safe


# ---------------------------------------------------------------------------
# XLSX → CSV conversion
# ---------------------------------------------------------------------------

def convert_xlsx_to_csv(xlsx_path: Path, out_dir: Path) -> list[dict]:
    """Convert each sheet of an XLSX to a separate CSV in out_dir.

    Returns list of {file, size_kb} for each sheet written.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    stem = xlsx_path.stem
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_sheet = "".join(c if c.isalnum() or c in "._-" else "_" for c in sheet_name)
        csv_name = f"{stem}_{safe_sheet}.csv"
        csv_path = out_dir / csv_name
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        results.append({"file": csv_name, "size_kb": csv_path.stat().st_size // 1024})
    wb.close()
    return results


# ---------------------------------------------------------------------------
# Readme
# ---------------------------------------------------------------------------

def write_readme(
    out_dir: Path,
    iso3: str,
    survey_year: float,
    dataset: dict,
    downloaded: list[dict],
    vintage: date | None,
) -> None:
    org = dataset.get("organization", {}).get("title", "unknown")
    hdx_url = f"https://data.humdata.org/dataset/{dataset.get('name', '')}"
    vintage_str = vintage.isoformat() if vintage else "unknown"
    gap_str = (
        f"{abs(vintage.year - int(survey_year))} year(s)"
        if vintage else "unknown"
    )
    # Admin-level counts embedded in the HDX notes field
    admin_lines = [
        l.strip()
        for l in dataset.get("notes", "").split("\n")
        if l.strip().startswith("- Admin")
    ]

    lines = [
        f"Country:          {iso3}",
        f"Survey year:      {survey_year:.4g}",
        f"Access date:      {ACCESS_DATE}",
        f"Source:           OCHA Humanitarian Data Exchange (HDX)",
        f"Dataset:          {dataset.get('title', '')}",
        f"HDX package:      cod-ab-{iso3.lower()}",
        f"Organization:     {org}",
        f"HDX URL:          {hdx_url}",
        f"Dataset date:     {dataset.get('dataset_date', '')}",
        f"Boundary vintage: {vintage_str}  (start of dataset_date range)",
        f"Gap to survey:    {gap_str}",
        "",
    ]
    if admin_lines:
        lines += ["Admin levels (from HDX dataset notes):"] + [f"  {l}" for l in admin_lines] + [""]
    lines += ["Files:"]
    for d in downloaded:
        lines.append(f"  {d['file']}  ({d['size_kb']} KB)")
        lines.append(f"    {d['url']}")
    lines += [
        "",
        "Note: 'boundary vintage' is the start date of the HDX dataset_date",
        "range, which OCHA uses to indicate when these boundaries first became",
        "valid for humanitarian use. It is distinct from the access date above.",
    ]
    (out_dir / "readme.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    survey_years = load_survey_years()
    print(f"Countries to process ({len(survey_years)}): {', '.join(sorted(survey_years))}\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest: dict = {"access_date": ACCESS_DATE, "countries": {}}

    for iso3 in sorted(survey_years):
        survey_year = survey_years[iso3]
        print(f"[{iso3}]  survey_year={survey_year:.4g}", end="  ", flush=True)

        dataset = fetch_cod_ab_package(iso3)
        time.sleep(REQUEST_DELAY)

        if dataset is None:
            print("NOT FOUND — no cod-ab package on HDX")
            manifest["countries"][iso3] = {"status": "not_found"}
            continue

        vintage = parse_dataset_date_start(dataset)
        resources = select_tabular_resources(dataset, survey_year)

        if not resources:
            print("dataset found but contains no tabular resources")
            manifest["countries"][iso3] = {
                "status": "no_tabular_resources",
                "dataset": dataset.get("name", ""),
            }
            continue

        vintage_str = vintage.isoformat() if vintage else "unknown"
        gap = abs(vintage.year - int(survey_year)) if vintage else "?"
        print(f"vintage={vintage_str}  gap={gap}yr  {len(resources)} file(s) to download")

        out_dir = OUTPUT_DIR / iso3
        out_dir.mkdir(parents=True, exist_ok=True)

        downloaded: list[dict] = []
        for res in resources:
            url = res.get("url", "")
            if not url:
                continue
            fname = safe_filename(res)
            dest = out_dir / fname
            is_xlsx = res.get("format", "").lower() in ("xlsx", "xls")

            # If already converted to CSV, collect the existing sheets and skip.
            if is_xlsx:
                existing_csvs = sorted(out_dir.glob(f"{dest.stem}_*.csv"))
                if existing_csvs:
                    print(f"  {fname}  already converted to CSV, skipping")
                    for cp in existing_csvs:
                        downloaded.append({"file": cp.name, "url": url, "size_kb": cp.stat().st_size // 1024})
                    continue

            if dest.exists():
                if is_xlsx:
                    # xlsx present but not yet converted — convert now.
                    try:
                        entries = convert_xlsx_to_csv(dest, out_dir)
                        dest.unlink()
                        for entry in entries:
                            print(f"  {entry['file']}  {entry['size_kb']} KB (converted from existing xlsx)")
                            downloaded.append({"file": entry["file"], "url": url, "size_kb": entry["size_kb"]})
                    except Exception as e:
                        print(f"  WARNING: could not convert existing {fname} to CSV: {e}")
                        downloaded.append({"file": fname, "url": url, "size_kb": dest.stat().st_size // 1024})
                else:
                    kb = dest.stat().st_size // 1024
                    print(f"  {fname}  already exists ({kb} KB), skipping")
                    downloaded.append({"file": fname, "url": url, "size_kb": kb})
                continue

            try:
                size = download_file(url, dest)
                kb = size // 1024
                if is_xlsx:
                    try:
                        entries = convert_xlsx_to_csv(dest, out_dir)
                        dest.unlink()
                        for entry in entries:
                            print(f"  {entry['file']}  {entry['size_kb']} KB")
                            downloaded.append({"file": entry["file"], "url": url, "size_kb": entry["size_kb"]})
                    except Exception as e:
                        print(f"  WARNING: could not convert {fname} to CSV: {e}")
                        downloaded.append({"file": fname, "url": url, "size_kb": kb})
                else:
                    print(f"  {fname}  {kb} KB")
                    downloaded.append({"file": fname, "url": url, "size_kb": kb})
                time.sleep(REQUEST_DELAY)
            except Exception as e:
                print(f"  ERROR downloading {fname}: {e}")

        if downloaded:
            write_readme(out_dir, iso3, survey_year, dataset, downloaded, vintage)
            manifest["countries"][iso3] = {
                "status": "ok",
                "dataset": dataset.get("name", ""),
                "boundary_vintage": vintage_str,
                "gap_years": gap,
                "files": [d["file"] for d in downloaded],
            }
        else:
            manifest["countries"][iso3] = {
                "status": "download_failed",
                "dataset": dataset.get("name", ""),
            }

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    counts: dict[str, int] = {}
    for v in manifest["countries"].values():
        counts[v["status"]] = counts.get(v["status"], 0) + 1

    print(f"\n{'='*52}")
    for status, n in sorted(counts.items()):
        print(f"  {status:<28} {n}")

    manifest_path = OUTPUT_DIR / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)
    print(f"\nManifest: {manifest_path}")


if __name__ == "__main__":
    main()
